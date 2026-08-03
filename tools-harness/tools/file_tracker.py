"""Excel tracker for inbox monitoring — supports PDF, Excel, and Word attachments."""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font

_COLUMNS = [
    "Timestamp", "Sender", "Subject", "Filename", "File Type",
    "File Path", "Converted Path", "Page Count", "Image Count", "File Size (KB)",
    "Summary", "Task Created", "Telegram Sent",
]

_ROW_KEYS = [
    "timestamp", "sender", "subject", "filename", "file_type",
    "file_path", "converted_path", "page_count", "image_count", "file_size_kb",
    "summary", "task_created", "telegram_sent",
]


def update_tracker(tracker_path: str, row: dict) -> None:
    """
    Update or create Excel tracking file with attachment processing log.

    Handles both new-style rows (file_type / file_path / converted_path) and
    legacy rows (pdf_path / md_path) by normalising keys before writing.

    Args:
        tracker_path: Absolute path to the Excel file (may not exist yet).
        row: Dict with keys matching _ROW_KEYS (legacy pdf_path/md_path also accepted).
    """
    # Normalise legacy key names from the old PDF-only schema
    row = dict(row)
    if "pdf_path" in row and "file_path" not in row:
        row["file_path"] = row.pop("pdf_path")
    if "md_path" in row and "converted_path" not in row:
        row["converted_path"] = row.pop("md_path")
    if "file_type" not in row:
        row["file_type"] = _infer_file_type(row.get("filename", ""))

    p = Path(tracker_path)
    p.parent.mkdir(parents=True, exist_ok=True)

    if p.exists():
        wb = openpyxl.load_workbook(str(p))
        ws = wb.active
        # If the existing sheet uses the old column layout, migrate the header
        if ws.cell(1, 5).value == "PDF Path":
            _migrate_header(ws)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inbox Attachment Log"
        ws.append(_COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    data_row = [row.get(k, "") for k in _ROW_KEYS]
    ws.append(data_row)
    wb.save(str(p))


def _infer_file_type(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    return {
        ".pdf": "pdf",
        ".xlsx": "excel",
        ".xls": "excel",
        ".docx": "word",
        ".doc": "word",
    }.get(ext, ext.lstrip(".") or "unknown")


def _migrate_header(ws) -> None:
    """Rewrite the header row of an old PDF-only tracker to the new schema."""
    for col_idx, col_name in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True)
