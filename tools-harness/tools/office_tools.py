"""
Office document processing utilities.

Extracts readable text from Excel (.xlsx/.xls) and Word (.docx/.doc) files
and writes a plain-text markdown summary alongside the source file.
Mirrors the (path, content) return contract of pdf_tools.pdf_to_markdown.
"""

from __future__ import annotations

from pathlib import Path


class PasswordProtectedError(Exception):
    """Raised when a file is encrypted/password-protected and cannot be read."""


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def excel_to_markdown(file_path: str, output_dir: str | None = None) -> tuple[str, str]:
    """
    Convert an Excel workbook to a markdown document.

    Each sheet becomes a level-2 heading.  Up to 200 rows × 20 columns
    per sheet are included to keep the summary prompt manageable.

    Returns:
        (md_file_path_str, md_content_str)
    """
    import openpyxl

    p = Path(file_path)
    if not p.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    out_dir = Path(output_dir) if output_dir is not None else p.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception as exc:
        msg = str(exc).lower()
        if "file is not a zip" in msg or "encrypted" in msg or "bad magic" in msg or "password" in msg:
            raise PasswordProtectedError(f"Excel file is password-protected: {p.name}") from exc
        raise
    parts: list[str] = [f"# {p.stem}\n"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"\n## {sheet_name}\n")
        rows_written = 0
        for row in ws.iter_rows(max_row=200, max_col=20, values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                parts.append("| " + " | ".join(cells) + " |")
                rows_written += 1
        if rows_written == 0:
            parts.append("_(empty sheet)_")

    wb.close()
    md_content = "\n".join(parts)
    md_file = out_dir / f"{p.stem}.md"
    md_file.write_text(md_content, encoding="utf-8")
    return (str(md_file), md_content)


def excel_metadata(file_path: str) -> dict:
    """Return sheet names and row counts without reading cell values."""
    import openpyxl

    p = Path(file_path)
    wb = openpyxl.load_workbook(str(p), read_only=True)
    meta = {}
    for name in wb.sheetnames:
        ws = wb[name]
        meta[name] = ws.max_row or 0
    wb.close()
    return meta


# ---------------------------------------------------------------------------
# Word
# ---------------------------------------------------------------------------

def docx_to_markdown(file_path: str, output_dir: str | None = None) -> tuple[str, str]:
    """
    Convert a Word document (.docx) to a markdown document.

    Headings map to markdown # levels; normal paragraphs are plain text.
    Tables are rendered as pipe-delimited rows.

    Returns:
        (md_file_path_str, md_content_str)
    """
    import docx as _docx

    p = Path(file_path)
    if not p.exists():
        raise FileNotFoundError(f"Word file not found: {file_path}")

    out_dir = Path(output_dir) if output_dir is not None else p.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        doc = _docx.Document(str(p))
    except Exception as exc:
        msg = str(exc).lower()
        if "encrypted" in msg or "password" in msg or "bad magic" in msg or "not a zip" in msg:
            raise PasswordProtectedError(f"Word file is password-protected: {p.name}") from exc
        raise
    parts: list[str] = [f"# {p.stem}\n"]

    for block in _iter_blocks(doc):
        parts.append(block)

    md_content = "\n".join(parts)
    md_file = out_dir / f"{p.stem}.md"
    md_file.write_text(md_content, encoding="utf-8")
    return (str(md_file), md_content)


def _iter_blocks(doc) -> list[str]:
    """Yield markdown-formatted strings for paragraphs and tables in order."""
    import docx as _docx
    from docx.oxml.ns import qn

    blocks: list[str] = []
    body = doc.element.body

    for child in body:
        tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
        if tag == "p":
            para = _docx.text.paragraph.Paragraph(child, doc)
            text = para.text.strip()
            if not text:
                continue
            style = para.style.name if para.style else ""
            if style.startswith("Heading 1"):
                blocks.append(f"\n# {text}")
            elif style.startswith("Heading 2"):
                blocks.append(f"\n## {text}")
            elif style.startswith("Heading 3"):
                blocks.append(f"\n### {text}")
            else:
                blocks.append(text)
        elif tag == "tbl":
            table = _docx.table.Table(child, doc)
            for row in table.rows:
                cells = [c.text.strip() for c in row.cells]
                blocks.append("| " + " | ".join(cells) + " |")

    return blocks


def docx_page_count(file_path: str) -> int:
    """
    Approximate page count from the Word document's built-in page count property.
    Returns 0 if not available.
    """
    import docx as _docx

    try:
        doc = _docx.Document(file_path)
        props = doc.core_properties
        # python-docx doesn't expose page count directly; fall back to 0
        _ = props.title  # just to confirm the file opens
        return 0
    except Exception:
        return 0
