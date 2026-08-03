"""
Edge case tests for tools/file_tracker.py.
Covers new schema, legacy key migration, missing keys, and header migration.
"""
import os
import sys
import openpyxl
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'tools-harness'))


def _new_row(**overrides):
    base = {
        "timestamp": "2026-04-10 09:00:00",
        "sender": "alice@example.com",
        "subject": "Q1 Report",
        "filename": "report.pdf",
        "file_type": "pdf",
        "file_path": "/agent/report.pdf",
        "converted_path": "/agent/report.md",
        "page_count": 3,
        "image_count": 1,
        "file_size_kb": 200,
        "summary": "Quarterly summary",
        "task_created": False,
        "telegram_sent": True,
    }
    base.update(overrides)
    return base


# ── New schema ──────────────────────────────────────────────────────────────────

def test_tracker_new_schema_creates_file(tmp_path):
    from tools.file_tracker import update_tracker
    p = tmp_path / "tracker.xlsx"
    update_tracker(str(p), _new_row())
    assert p.exists()
    wb = openpyxl.load_workbook(str(p))
    ws = wb.active
    assert ws.max_row == 2  # header + 1 data row


def test_tracker_new_schema_file_type_column(tmp_path):
    from tools.file_tracker import update_tracker
    p = tmp_path / "tracker.xlsx"
    update_tracker(str(p), _new_row(filename="data.xlsx", file_type="excel"))
    wb = openpyxl.load_workbook(str(p))
    ws = wb.active
    # File Type is column 5
    assert ws.cell(row=1, column=5).value == "File Type"
    assert ws.cell(row=2, column=5).value == "excel"


def test_tracker_appends_multiple_rows(tmp_path):
    from tools.file_tracker import update_tracker
    p = tmp_path / "tracker.xlsx"
    for ext, ft in [("a.pdf", "pdf"), ("b.xlsx", "excel"), ("c.docx", "word")]:
        update_tracker(str(p), _new_row(filename=ext, file_type=ft))
    wb = openpyxl.load_workbook(str(p))
    assert wb.active.max_row == 4  # header + 3


# ── Legacy key migration ────────────────────────────────────────────────────────

def test_tracker_legacy_keys_normalised(tmp_path):
    """pdf_path/md_path from old rows must be stored under file_path/converted_path."""
    from tools.file_tracker import update_tracker
    p = tmp_path / "legacy.xlsx"
    legacy_row = {
        "timestamp": "2026-04-10",
        "sender": "old@example.com",
        "subject": "Old email",
        "filename": "old.pdf",
        "pdf_path": "/agent/old.pdf",
        "md_path": "/agent/old.md",
        "page_count": 2,
        "image_count": 0,
        "file_size_kb": 50,
        "summary": "Old summary",
        "task_created": False,
        "telegram_sent": False,
    }
    update_tracker(str(p), legacy_row)
    wb = openpyxl.load_workbook(str(p))
    ws = wb.active
    # file_path is column 6, converted_path is column 7
    assert ws.cell(row=2, column=6).value == "/agent/old.pdf"
    assert ws.cell(row=2, column=7).value == "/agent/old.md"


def test_tracker_legacy_keys_do_not_overwrite_new_keys(tmp_path):
    """If both pdf_path and file_path are present, file_path wins."""
    from tools.file_tracker import update_tracker
    p = tmp_path / "both.xlsx"
    row = _new_row(file_path="/new/path.pdf")
    row["pdf_path"] = "/old/path.pdf"  # should be ignored
    update_tracker(str(p), row)
    wb = openpyxl.load_workbook(str(p))
    assert wb.active.cell(row=2, column=6).value == "/new/path.pdf"


# ── file_type inference ─────────────────────────────────────────────────────────

def test_infer_file_type_all_extensions(tmp_path):
    from tools.file_tracker import _infer_file_type
    assert _infer_file_type("report.pdf") == "pdf"
    assert _infer_file_type("data.xlsx") == "excel"
    assert _infer_file_type("data.xls") == "excel"
    assert _infer_file_type("memo.docx") == "word"
    assert _infer_file_type("memo.doc") == "word"
    assert _infer_file_type("archive.zip") == "zip"
    assert _infer_file_type("noextension") == "unknown"


def test_infer_file_type_uppercase_extension(tmp_path):
    """Extensions like .PDF or .XLSX must be normalised."""
    from tools.file_tracker import _infer_file_type
    assert _infer_file_type("REPORT.PDF") == "pdf"
    assert _infer_file_type("DATA.XLSX") == "excel"


# ── Missing / partial keys ──────────────────────────────────────────────────────

def test_tracker_missing_keys_do_not_crash(tmp_path):
    """A row with only a sender should not raise — missing fields become empty."""
    from tools.file_tracker import update_tracker
    p = tmp_path / "sparse.xlsx"
    update_tracker(str(p), {"sender": "sparse@example.com"})
    assert p.exists()


def test_tracker_creates_nested_directory(tmp_path):
    from tools.file_tracker import update_tracker
    nested = tmp_path / "a" / "b" / "c" / "tracker.xlsx"
    update_tracker(str(nested), _new_row())
    assert nested.exists()


# ── Header migration ────────────────────────────────────────────────────────────

def test_tracker_migrates_old_header(tmp_path):
    """An existing tracker with old 'PDF Path' header gets migrated on next write."""
    from tools.file_tracker import update_tracker
    import openpyxl
    from openpyxl.styles import Font

    p = tmp_path / "old_tracker.xlsx"

    # Build a file with the old column layout (PDF Path at col 5)
    old_cols = [
        "Timestamp", "Sender", "Subject", "Filename", "PDF Path",
        "MD Path", "Page Count", "Image Count", "File Size (KB)",
        "Summary", "Task Created", "Telegram Sent",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF Inbox Log"
    ws.append(old_cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(str(p))

    # Writing a new row should migrate the header
    update_tracker(str(p), _new_row())
    wb2 = openpyxl.load_workbook(str(p))
    ws2 = wb2.active
    assert ws2.cell(row=1, column=5).value == "File Type"
    assert ws2.cell(row=1, column=6).value == "File Path"
